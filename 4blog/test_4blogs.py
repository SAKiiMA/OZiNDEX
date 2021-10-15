import pytest
import os
from pathlib import Path
import openpyxl
import pandas as pd

from utils_4blog import(
    load_data,
    get_transposed_df,
    get_clean_df,
    resolve_country_name, 
    get_merged_df)


EXCEL_PATH = Path("2019-20-historical-migration-statistics-locked.xlsx")


def test_excel_exist():
    #make sure excel file exists
    assert os.path.isfile(EXCEL_PATH)


def test_excel_sheet():
    # make sure sheets 3.2 & 3.3 exist
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    assert '3.2' in wb.sheetnames
    assert '3.3' in wb.sheetnames

    
@pytest.fixture
def get_transpose():
    # load excel sheets in test environment
    sh1, sh2 = load_data()
    dft1 = get_transposed_df(sh1)
    dft2 = get_transposed_df(sh2)
    return dft1, dft2


def test_transposed_df(get_transpose):
    # make sure columns for transposed df is correct
    dft1, dft2 = get_transpose
    assert len(dft1.columns) == 23    
    assert len(dft2.columns) == 5
 

@pytest.fixture
def get_clean(get_transpose):
    dft1, dft2 = get_transpose
    dfc1 = get_clean_df(dft1)
    dfc2 = get_clean_df(dft2)
    return dfc1, dfc2
    
    
@pytest.mark.filterwarnings("ignore")
def test_clean_df(get_clean):
    # make sure cleaned data frame are in right shape
    dfc1, dfc2 = get_clean
    assert "Stream" in dfc1.columns
    assert "Country" in dfc1.columns
    assert "1996–97" in dfc1.columns
    assert "1999–00" in dfc1.columns
    assert "2016–17" in dfc1.columns
    assert "Stream" in dfc2.columns
    assert "Country" in dfc2.columns
    assert "2017–18" in dfc2.columns
    assert "2019–20" in dfc2.columns
    assert len(dfc1.columns) == 23
    assert len(dfc2.columns) == 5
    assert dfc1["Stream"].dtype == object
    assert dfc1["2005–06"].dtype == int
    assert dfc1["2011–12"].dtype == int
    assert dfc2["Country"].dtype == object
    assert dfc2["2018–19"].dtype == int
  

@pytest.fixture
def get_resolved(get_clean):
    dfc1, dfc2 = get_clean
    df1, df2 = resolve_country_name(dfc1, dfc2)
    return df1, df2


@pytest.mark.filterwarnings("ignore")
def test_resolved(get_resolved):
    df1, df2 = get_resolved
    assert len(df1['Country'].unique()) == len(df2['Country'].unique())
    
    
@pytest.fixture
def get_merged(get_resolved):
    df1, df2 = get_resolved
    df = get_merged_df(df1, df2)
    return df


@pytest.mark.filterwarnings("ignore")
def test_merged(get_resolved):
    df1, df2 = get_resolved
    df = get_merged_df(df1, df2)
    assert len(df.columns) == 26
    assert df["2005–06"].dtype == int
    assert df["2011–12"].dtype == int
    assert df["2018–19"].dtype == int
    
    assert df[df["Country"] == "Iran"].sum(numeric_only=True).sum(
    ) == df1[df1["Country"] == "Iran"].sum(numeric_only=True).sum(
    ) + df2[df2["Country"] == "Iran"].sum(numeric_only=True).sum()
    
    assert df[df["Country"] == "India"].sum(numeric_only=True).sum(
    ) == df1[df1["Country"] == "India"].sum(numeric_only=True).sum(
    ) + df2[df2["Country"] == "India"].sum(numeric_only=True).sum()
    
    assert df[(df["Country"] == "France") & (df["Stream"] == "Skill stream")
             ]["2011–12"].values == df1[(df1["Country"] == "France") & 
                                        (df1["Stream"] == "Skill stream")]["2011–12"].values
    
    assert df[(df["Country"] == "Italy") & (df["Stream"] == "Special Eligibility")
             ]["2018–19"].values == df2[(df2["Country"] == "Italy") & 
                                        (df2["Stream"] == "Special Eligibility")]["2018–19"].values

